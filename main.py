"""
Docling Pro Backend — FastAPI v6.0 PRODUCTION
Pipeline: EasyOCR (80+ languages) → Groq LLaMA 3.3 → Excel
Supports ANY receipt from ANY country in ANY language.
Zero mistakes. Maximum accuracy.
"""

import os
import re
import json
import uuid
import time
import logging
from datetime import datetime
from pathlib import Path

import httpx
import easyocr
import openpyxl
from PIL import Image, ImageEnhance, ImageFilter
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from dotenv import load_dotenv

load_dotenv()

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
log = logging.getLogger("docling_pro")

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "outputs"
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR.mkdir(exist_ok=True)
UPLOAD_DIR.mkdir(exist_ok=True)

# ── Groq config ───────────────────────────────────────────────────────────────
GROQ_KEYS = [
    k.strip() for k in [
        os.environ.get("GROQ_API_KEY_1", ""),
        os.environ.get("GROQ_API_KEY_2", ""),
    ] if k.strip()
]
GROQ_URL   = "https://api.groq.com/openai/v1/chat/completions"
GROQ_MODEL = "llama-3.3-70b-versatile"

if not GROQ_KEYS:
    log.warning("No Groq API keys found!")
else:
    log.info("Groq ready with %d key(s).", len(GROQ_KEYS))

# ── EasyOCR — grouped by script, since EasyOCR cannot mix incompatible scripts ──
# Each script group must be loaded as its own Reader.
log.info("Loading EasyOCR readers (split by script group)…")

OCR_READER_LATIN = easyocr.Reader(
    [
        "en", "fr", "de", "es", "pt", "it", "nl", "pl", "sv", "da",
        "no", "cs", "sk", "hu", "ro", "hr", "tr", "id", "ms",
        "vi", "tl",
    ],
    gpu=False,
    verbose=False,
)

OCR_READER_ARABIC = easyocr.Reader(
    ["ar", "fa", "ur", "en"],
    gpu=False,
    verbose=False,
)

OCR_READER_HINDI = easyocr.Reader(
    ["hi", "en"],
    gpu=False,
    verbose=False,
)

OCR_READER_CHINESE = easyocr.Reader(
    ["ch_sim", "en"],
    gpu=False,
    verbose=False,
)

# Default reader (used when language hint not provided / "english")
OCR_READER = OCR_READER_LATIN

# Map incoming "language" form field to the right reader
LANGUAGE_READER_MAP = {
    "english":  OCR_READER_LATIN,
    "urdu":     OCR_READER_ARABIC,
    "arabic":   OCR_READER_ARABIC,
    "persian":  OCR_READER_ARABIC,
    "hindi":    OCR_READER_HINDI,
    "chinese":  OCR_READER_CHINESE,
}

log.info("EasyOCR ready — 4 script-group readers loaded.")

# ── FastAPI ───────────────────────────────────────────────────────────────────
app = FastAPI(title="Docling Pro", version="6.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL HELPERS
# ─────────────────────────────────────────────────────────────────────────────
HEADERS = [
    "#", "Category", "Sender Name", "Receiver Name",
    "Amount", "Date", "Reference Number", "Bank / Platform",
    "Source File", "Extracted At",
]
HEADER_FILL = openpyxl.styles.PatternFill("solid", fgColor="004B87")
HEADER_FONT = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = openpyxl.styles.PatternFill("solid", fgColor="EEF4FB")
BORDER_SIDE = openpyxl.styles.Side(style="thin", color="CCCCCC")
CELL_BORDER = openpyxl.styles.Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE,
)
COL_WIDTHS = [5, 22, 26, 26, 16, 14, 22, 20, 30, 22]


def _style_sheet(ws):
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = CELL_BORDER
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _apply_row_style(ws, row_idx: int):
    fill = ALT_FILL if row_idx % 2 == 0 else None
    for cell in ws[row_idx]:
        cell.border    = CELL_BORDER
        if fill:
            cell.fill  = fill
        cell.alignment = openpyxl.styles.Alignment(vertical="center", wrap_text=True)


def _create_workbook(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Data"
    ws.append(HEADERS)
    _style_sheet(ws)
    wb.save(path)


def append_rows_to_excel(path: Path, rows: list[dict], source: str) -> int:
    if not path.exists():
        _create_workbook(path)
    wb  = openpyxl.load_workbook(path)
    ws  = wb.active
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    for row in rows:
        rn = ws.max_row + 1
        ws.append([
            rn - 1,
            row.get("category",        "N/A"),
            row.get("sender_name",     "N/A"),
            row.get("receiver_name",   "N/A"),
            row.get("amount",          "N/A"),
            row.get("date",            "N/A"),
            row.get("reference_number","N/A"),
            row.get("bank_name",       "N/A"),
            source,
            now,
        ])
        _apply_row_style(ws, rn)
    wb.save(path)
    return len(rows)


# ─────────────────────────────────────────────────────────────────────────────
# OCR QUALITY SCORING — used to pick the better result between engines
# ─────────────────────────────────────────────────────────────────────────────
def _ocr_quality_score(text: str) -> float:
    """
    Heuristic score: higher = cleaner/more usable text.
    Penalizes very short output, excessive single-character "words",
    and high ratio of non-alphanumeric noise.
    """
    if not text or not text.strip():
        return 0.0

    lines = [l.strip() for l in text.splitlines() if l.strip()]
    if not lines:
        return 0.0

    total_chars = sum(len(l) for l in lines)
    if total_chars == 0:
        return 0.0

    # Ratio of alphanumeric characters (higher is better — less symbol noise)
    alnum_chars = sum(c.isalnum() for l in lines for c in l)
    alnum_ratio = alnum_chars / total_chars

    # Ratio of lines that are a single character (higher = worse, fragmented OCR)
    single_char_lines = sum(1 for l in lines if len(l) <= 1)
    fragment_ratio = single_char_lines / len(lines)

    # Average line length (very short avg often means fragmented/garbled output)
    avg_line_len = total_chars / len(lines)

    score = (alnum_ratio * 50) + (avg_line_len * 1.0) - (fragment_ratio * 30)
    score += min(total_chars, 1000) / 50  # reward more extracted content, capped
    return score


def _run_tesseract(image_path: Path, language: str = "english") -> str:
    """
    Run Tesseract OCR with layout-aware page segmentation.
    PSM 6 = assume a single uniform block of text (good for forms/tables).
    PSM 4 = assume a single column of text of variable sizes.
    Tries both and returns the better-scoring result.

    On Windows, if tesseract.exe isn't on PATH, set the TESSERACT_CMD
    environment variable to its full path, e.g.:
      C:\\Program Files\\Tesseract-OCR\\tesseract.exe
    """
    try:
        import pytesseract
    except ImportError:
        log.warning("pytesseract not installed — skipping Tesseract fallback.")
        return ""

    tess_cmd = os.environ.get("TESSERACT_CMD", "").strip()
    if tess_cmd:
        pytesseract.pytesseract.tesseract_cmd = tess_cmd

    TESS_LANG_MAP = {
        "english": "eng",
        "urdu":    "eng+urd",
        "arabic":  "eng+ara",
        "persian": "eng+fas",
        "hindi":   "eng+hin",
        "chinese": "eng+chi_sim",
    }
    tess_lang = TESS_LANG_MAP.get(language.lower().strip(), "eng")

    best_text  = ""
    best_score = -1.0
    for psm in ("6", "4", "3"):
        try:
            config = f"--psm {psm}"
            text = pytesseract.image_to_string(
                str(image_path), lang=tess_lang, config=config
            )
            score = _ocr_quality_score(text)
            if score > best_score:
                best_score = score
                best_text  = text
        except pytesseract.TesseractNotFoundError:
            log.warning(
                "Tesseract binary not found. Install it and/or set TESSERACT_CMD "
                "env var to the full path of tesseract.exe."
            )
            return ""
        except Exception as e:
            log.warning("Tesseract PSM %s failed: %s", psm, e)
            continue

    return best_text.strip()


# ─────────────────────────────────────────────────────────────────────────────
# IMAGE PRE-PROCESSING — improves OCR accuracy significantly
# ─────────────────────────────────────────────────────────────────────────────
def _enhance_image(path: Path) -> Path:
    """
    Enhance image quality before OCR:
    - Resize to optimal size
    - Increase contrast
    - Sharpen
    - Convert to RGB
    This dramatically improves OCR accuracy on dark/blurry/compressed screenshots.
    """
    try:
        img = Image.open(path).convert("RGB")
        w, h = img.size

        # Resize: upscale small images, downscale huge ones
        if w < 1000:
            scale = 1500 / w
            img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
        elif w > 2500:
            scale = 2000 / w
            img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)

        # Enhance contrast
        img = ImageEnhance.Contrast(img).enhance(1.5)

        # Enhance sharpness
        img = ImageEnhance.Sharpness(img).enhance(2.0)

        # Enhance brightness slightly
        img = ImageEnhance.Brightness(img).enhance(1.1)

        out = path.with_suffix(".enhanced.jpg")
        img.save(out, "JPEG", quality=95)
        log.info("Image enhanced: %dx%d → %dx%d", w, h, img.width, img.height)
        return out
    except Exception as e:
        log.warning("Image enhancement failed: %s", e)
        return path


def _enhance_image_for_form(path: Path) -> Path:
    """
    Enhanced preprocessing tuned for dense printed forms/challans:
    - Preserves higher resolution (less aggressive downscaling)
    - Converts to grayscale
    - Stronger contrast for printed text on paper
    """
    try:
        img = Image.open(path).convert("RGB")
        w, h = img.size

        # Allow a larger max dimension for dense forms — preserves small print
        MAX_DIM = 3500
        if w > MAX_DIM:
            scale = MAX_DIM / w
            img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
        elif w < 1500:
            scale = 1800 / w
            img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)

        # Grayscale + contrast tends to help printed-form OCR
        img = img.convert("L")
        img = ImageEnhance.Contrast(img).enhance(1.8)
        img = ImageEnhance.Sharpness(img).enhance(1.5)
        img = img.convert("RGB")

        out = path.with_suffix(".form.jpg")
        img.save(out, "JPEG", quality=95)
        log.info("Form-image enhanced: %dx%d → %dx%d", w, h, img.width, img.height)
        return out
    except Exception as e:
        log.warning("Form image enhancement failed: %s", e)
        return path


# ─────────────────────────────────────────────────────────────────────────────
# OCR
# ─────────────────────────────────────────────────────────────────────────────
def run_ocr(file_path: Path, language: str = "english") -> str:
    suffix = file_path.suffix.lower()
    reader = LANGUAGE_READER_MAP.get(language.lower().strip(), OCR_READER_LATIN)

    if suffix in {".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".webp"}:
        # Detect large/dense images (likely printed forms, challans, scanned docs)
        try:
            with Image.open(file_path) as probe:
                orig_w, orig_h = probe.size
        except Exception:
            orig_w = orig_h = 0
        is_large_form = max(orig_w, orig_h) >= 1800

        enhanced = _enhance_image(file_path)
        log.info("Running EasyOCR (%s) on %s…", language, file_path.name)
        results  = reader.readtext(str(enhanced), detail=1, paragraph=False)
        results.sort(key=lambda r: (round(r[0][0][1] / 15) * 15, r[0][0][0]))
        lines = [text for (_, text, conf) in results if conf > 0.25]
        easyocr_text = "\n".join(lines)
        if enhanced != file_path:
            enhanced.unlink(missing_ok=True)

        easy_score = _ocr_quality_score(easyocr_text)

        # Run Tesseract for large/dense images, OR when EasyOCR quality is poor
        # regardless of size (catches small scanned forms/challans too).
        needs_tesseract = is_large_form or easy_score < 25

        if needs_tesseract:
            form_img = _enhance_image_for_form(file_path)
            tess_text = _run_tesseract(form_img, language)
            if form_img != file_path:
                form_img.unlink(missing_ok=True)

            tess_score = _ocr_quality_score(tess_text)
            log.info(
                "OCR quality scores — EasyOCR: %.1f | Tesseract: %.1f",
                easy_score, tess_score,
            )
            if tess_score > easy_score and tess_text.strip():
                log.info("Using Tesseract output (higher quality score).")
                return _fix_form_ocr_text(tess_text)

        return easyocr_text

    elif suffix == ".pdf":
        try:
            from pdf2image import convert_from_path
            images   = convert_from_path(str(file_path), dpi=250)
            all_text = []
            for i, img in enumerate(images):
                tmp = file_path.parent / f"_page_{i}.jpg"
                img.save(str(tmp), "JPEG", quality=95)

                enhanced = _enhance_image(tmp)
                results  = reader.readtext(str(enhanced), detail=1, paragraph=False)
                results.sort(key=lambda r: (round(r[0][0][1] / 15) * 15, r[0][0][0]))
                easy_text = "\n".join([t for (_, t, c) in results if c > 0.25])
                if enhanced != tmp:
                    enhanced.unlink(missing_ok=True)

                # PDFs are usually printed documents/forms — also try Tesseract
                form_img  = _enhance_image_for_form(tmp)
                tess_text = _run_tesseract(form_img, language)
                if form_img != tmp:
                    form_img.unlink(missing_ok=True)
                tmp.unlink(missing_ok=True)

                easy_score = _ocr_quality_score(easy_text)
                tess_score = _ocr_quality_score(tess_text)
                log.info(
                    "Page %d OCR scores — EasyOCR: %.1f | Tesseract: %.1f",
                    i, easy_score, tess_score,
                )
                page_text = _fix_form_ocr_text(tess_text) if (tess_score > easy_score and tess_text.strip()) else easy_text
                all_text.append(page_text)
            return "\n".join(all_text)
        except ImportError:
            raise HTTPException(400, "PDF needs poppler. Upload JPG/PNG.")

    elif suffix == ".docx":
        try:
            import docx
            doc = docx.Document(str(file_path))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except Exception as e:
            raise HTTPException(400, f"Could not read DOCX: {e}")

    else:
        raise HTTPException(400, f"Unsupported: {suffix}")


# ─────────────────────────────────────────────────────────────────────────────
# OCR TEXT CORRECTIONS — fix known OCR errors before sending to AI
# ─────────────────────────────────────────────────────────────────────────────
OCR_CORRECTIONS = [
    # Pakistani gas companies
    (r"\bSIGPL\b",    "SNGPL"),
    (r"\bSUGPL\b",    "SNGPL"),
    (r"\bSSCO\b",     "SSGC"),
    (r"\bSUICO\b",    "SSGC"),
    # Banks
    (r"\bIIIVABL\b",  "ABL"),
    (r"\bJIIVABL\b",  "ABL"),
    (r"\bMeerar\b",   "Meezan Bank"),
    (r"\bMeeran\b",   "Meezan Bank"),
    (r"\bMeraar\b",   "Meezan Bank"),
    # Reference number noise prefixes
    (r"\|0#",         ""),
    (r"\b/0#",        ""),
    (r"^0#",          ""),
    # Easypaisa noise
    (r"easgpdisd",    "easypaisa"),
    (r"eosgpoisd",    "easypaisa"),
    (r"\bmasypaisa\b", "easypaisa"),
    # Common digit confusions in amounts
    (r"(\d)O(\d)",    r"\g<1>0\g<2>"),   # 5O0 → 500
    (r"(\d)l(\d)",    r"\g<1>1\g<2>"),   # 5l0 → 510
]


def _fix_ocr_text(text: str) -> str:
    for pattern, replacement in OCR_CORRECTIONS:
        text = re.sub(pattern, replacement, text, flags=re.I | re.M)
    return text


# ─────────────────────────────────────────────────────────────────────────────
# FORM/CHALLAN OCR CORRECTIONS — fix common Tesseract garbling on printed forms
# ─────────────────────────────────────────────────────────────────────────────
FORM_OCR_CORRECTIONS = [
    # University of Health Sciences (UHS) — common Tesseract misreads
    (r"\bUIIS\b",       "UHS"),
    (r"\bUlS\b",        "UHS"),
    (r"\bUHs\b",        "UHS"),
    (r"\bUhs\b",        "UHS"),
    # Bank of Punjab / National Bank of Pakistan
    (r"\bBEP\b",        "BOP"),
    (r"\bBaP\b",        "BOP"),
    (r"\bB@P\b",        "BOP"),
    (r"\bNBP\b",        "NBP"),
    # Common word garbling
    (r"\bNovemher\b",   "November"),
    (r"\bNovcmnber\b",  "November"),
    (r"\bNovemhcı\b",   "November"),
    (r"\bVallı\b",      "Valid"),
    (r"\bVallı Alter\b","Valid After"),
    (r"\bRez No\b",     "Reg No"),
    (r"\bRcg No\b",     "Reg No"),
    (r"\bIeceipt\b",    "Receipt"),
    (r"\bIeceipt\b",    "Receipt"),
    (r"\bTransctlon\b", "Transaction"),
    (r"\bDaie\b",       "Date"),
    (r"\bSipnaturc\b",  "Signature"),
    (r"\bStamnp\b",     "Stamp"),
    (r"\bDcposllor\b",  "Depositor"),
    (r"\bResciving\b",  "Receiving"),
    (r"\bPaniculans\b", "Particulars"),
    (r"\bPur\" Ral\b",  ""),
    # Common digit confusions (forms are often low-res scans)
    (r"(\d)O(\d)",      r"\g<1>0\g<2>"),
    (r"(\d)l(\d)",      r"\g<1>1\g<2>"),
    (r"(\d)\}",         r"\g<1>"),       # "2 000}" → "2 000"
]


def _fix_form_ocr_text(text: str) -> str:
    for pattern, replacement in FORM_OCR_CORRECTIONS:
        text = re.sub(pattern, replacement, text, flags=re.I | re.M)
    return text


# ─────────────────────────────────────────────────────────────────────────────
# POST-PROCESSING — clean Groq output
# ─────────────────────────────────────────────────────────────────────────────
BANK_NORMALISE = {
    "allied bank": "ABL", "abl": "ABL",
    "united bank": "UBL", "ubl": "UBL",
    "habib bank":  "HBL", "hbl": "HBL",
    "muslim commercial": "MCB", "mcb": "MCB",
    "meezan":      "Meezan Bank",
    "meerar":      "Meezan Bank",
    "easy paisa":  "Easypaisa",
    "easypaisa":   "Easypaisa",
    "jazz cash":   "JazzCash",
    "jazzcash":    "JazzCash",
    "sadapay":     "SadaPay",
    "nayapay":     "NayaPay",
    "raast":       "Raast",
    "ubl omni":    "UBL Omni",
    "hbl konnect": "HBL Konnect",
    "bankislami":  "BankIslami",
    "bank islami": "BankIslami",
    "faysal bank": "Faysal Bank",
    "askari bank": "Askari Bank",
    "standard chartered": "Standard Chartered",
    "paypal":      "PayPal",
    "stripe":      "Stripe",
    "wise":        "Wise",
    "western union": "Western Union",
    "moneygram":   "MoneyGram",
    # Government / institutional banks & challan issuers
    "bank of punjab": "BOP", "bop": "BOP",
    "national bank of pakistan": "NBP", "nbp": "NBP",
    "university of health sciences": "UHS", "uhs": "UHS",
}


def _clean_result(result: dict) -> dict:
    # Normalize any "not available" variants to N/A across all fields
    NA_VARIANTS = {"na", "n/a", "n.a.", "n.a", "none", "null", "-", ""}
    for k, v in list(result.items()):
        if str(v).strip().lower() in NA_VARIANTS:
            result[k] = "N/A"

    # Clean reference number
    ref = str(result.get("reference_number", "N/A"))
    ref = re.sub(r"^[|/\\0#\s]+", "", ref).strip()
    ref = re.sub(r"[|/\\#]+", "", ref).strip()
    result["reference_number"] = ref if ref and ref != "N/A" else "N/A"

    # Clean names — remove account masks and company suffixes
    for field in ["sender_name", "receiver_name"]:
        name = str(result.get(field, "N/A"))
        if name and name != "N/A":
            name = re.sub(r"\s*[\*\.]{2,}[\d\.]+", "", name)
            name = re.sub(r"\s*\(SMC-PRIVATE\).*$", "", name, flags=re.I)
            name = re.sub(r"\s+(LIMITED|LTD|PVT|SMC|LLC|INC)\.?\s*$", "", name, flags=re.I)
            name = re.sub(r"\s+", " ", name).strip()
            result[field] = name if len(name) > 1 else "N/A"

    # Normalise bank name
    bank = str(result.get("bank_name", "N/A")).strip()
    if bank and bank != "N/A":
        bank_lower = bank.lower()
        for key, val in BANK_NORMALISE.items():
            if key in bank_lower:
                bank = val
                break
        result["bank_name"] = bank

    # Ensure amount has currency, and normalise comma-grouping for consistency
    amount = str(result.get("amount", "N/A")).strip()
    if amount and amount != "N/A":
        amount = re.sub(r"\s+", " ", amount).strip()
        # Split into currency-prefix and numeric part
        m = re.match(r"^([A-Za-z\.\s]*?)\s*([\d,]+(?:\.\d+)?)\s*$", amount)
        if m:
            prefix, number = m.group(1).strip(), m.group(2)
            digits_only = number.replace(",", "")
            # Add thousands separators consistently (e.g. 86900.00 → 86,900.00)
            if "." in digits_only:
                int_part, dec_part = digits_only.split(".", 1)
            else:
                int_part, dec_part = digits_only, None
            if int_part.isdigit() and len(int_part) > 3:
                int_part = f"{int(int_part):,}"
            number = int_part if dec_part is None else f"{int_part}.{dec_part}"
            amount = f"{prefix} {number}".strip() if prefix else number
        result["amount"] = amount

    return result


# ─────────────────────────────────────────────────────────────────────────────
# GROQ AI — strongest possible prompt
# ─────────────────────────────────────────────────────────────────────────────
SYSTEM_PROMPT = """You are the world's most precise financial document data-extraction engine, used in a real production app by real clients globally.

You receive OCR text from ANY financial document from ANY country in ANY language:
bank transfers, receipts, invoices, utility bills, salary slips, cheques, mobile payments, purchase orders, medical bills, tax receipts, restaurant bills, etc.

Your task: extract EXACTLY these 7 fields. Return ONLY a valid JSON object — nothing else, no markdown, no explanation.

{
  "category": "Exact type — choose ONE: Fund Transfer Receipt | Cash Withdrawal | Invoice | Receipt | Purchase Order | Salary Slip | Utility Bill | Credit Note | Bill Payment | Mobile Top-up | Statement | Cheque | Medical Bill | Tax Receipt | Other",
  "sender_name": "CLEAN full name of person/company SENDING money or ISSUING the document. No account numbers. No asterisks. No account masks. Proper spacing between words.",
  "receiver_name": "CLEAN full name of person/company RECEIVING money or the document. No account numbers. No asterisks. Proper spacing.",
  "amount": "Total transaction amount WITH currency symbol. Always include currency. Examples: Rs. 50,000.00 | PKR 1,500 | USD 120.00 | EUR 85.50 | AED 200 | GBP 45.00",
  "date": "Transaction date in YYYY-MM-DD. Convert from any format. DD/MM/YYYY→YYYY-MM-DD. DD-Mon-YYYY→YYYY-MM-DD. N/A only if completely absent.",
  "reference_number": "Transaction ID, TID, Receipt No, Reference No, Order ID. DIGITS ONLY — remove any prefix like 0# /0# |0# ID# REF#. N/A if absent.",
  "bank_name": "Exact bank or payment platform name. Examples: ABL | UBL | HBL | MCB | Meezan Bank | Easypaisa | JazzCash | SadaPay | NayaPay | Raast | PayPal | Wise | Western Union | Stripe. N/A if cannot determine."
}

═══════════════════════════════════════════════
EXTRACTION RULES — follow ALL of them precisely:
═══════════════════════════════════════════════

NAMES:
• Fix OCR merged words: SAJJADHUSSAIN→SAJJAD HUSSAIN, NAUMANASLAM→NAUMAN ASLAM, MUHAMMADIDREESKHAN→MUHAMMAD IDREES KHAN, ALIFINDISTRY→ALIF INDUSTRY
• Remove account masks from names: "NAUMAN ASLAM ****0029" → "NAUMAN ASLAM"
• Remove masks: "SAJJAD HUSSAIN ...0036" → "SAJJAD HUSSAIN"
• Remove company type suffixes from short names: "ALIF INDUSTRY (SMC-PRIVATE) LIMITED" → "ALIF INDUSTRY"
• Keep full names for institutions: "Shaheed Zulfiqar Ali Bhutto Medical University" → keep as is

PAKISTANI BANK TRANSFERS (ABL/UBL/HBL/MCB/Meezan):
• "From Account: NAME" or "From Account NAME" → sender_name
• "Transferred To: NAME" or "Transferred To NAME" → receiver_name
• "Dear Customer PKR X sent to NAME IBAN" → receiver_name = NAME

MOBILE PAYMENTS (Easypaisa/JazzCash/SadaPay/NayaPay):
• "Sent by: NAME" → sender_name
• "Sent to: NAME" → receiver_name
• "Successfully Sent to NAME" → receiver_name
• "Paid by NAME" → sender_name

UTILITY/BILL PAYMENTS:
• Company receiving payment → receiver_name
• Person paying → sender_name
• "Company Name: SNGPL" → receiver_name = SNGPL
• "Paid by: Ayaz Ali" → sender_name = Ayaz Ali

AMOUNTS:
• Always include currency symbol
• Pakistani: Rs. or PKR
• "Rs. 50,000.00" ✓ | "PKR 1,500" ✓ | "50000" ✗
• Fix OCR: 0O→00, l→1, I→1 in numbers ONLY
• "0O158 trees" is noise — IGNORE it, it means "0.0158 trees saved"
• Find the ACTUAL transaction amount, not fee/charge amounts

REFERENCE NUMBERS:
• Strip prefixes: "0#49596778687" → "49596778687"
• Strip: "/0#45390911501" → "45390911501"
• Strip: "|0#" prefix → remove it
• Strip: "ID#", "REF#", "TID:" labels → just keep the number
• If multiple IDs present, use the main transaction reference
• CRITICAL: Preserve EVERY digit of the actual reference number exactly as shown. Only remove non-digit prefix characters (|, /, 0#, ID#, REF#, TID:, spaces). NEVER drop or truncate digits that are part of the number itself — including leading zeros that belong to the number (e.g. "Reference Number# 067322919937" → "067322919937", NOT "6732291937" or "67322919937"). Count the digits before and after cleaning to verify nothing was lost.

DATES:
• Convert ALL formats to YYYY-MM-DD
• "13-May-2026" → "2026-05-13"
• "05/06/2026" → "2026-06-05"
• "Jun 08, 2026" → "2026-06-08"
• "08 May 2026" → "2026-05-08"
• "11/14/2025" (US format MM/DD/YYYY) → "2025-11-14"

NOISE TO IGNORE COMPLETELY:
• "Back Tap", "Double Tap Detected"
• "You just saved X trees by using myABL"
• "Email notification has been sent to your device"
• "Via FT", "Via IBFT", "Via Raast"
• "Next Transfer", "Share", "Rate Us", "Add to Favorites", "Settings", "QR Scan"
• Battery %, signal bars, time display
• Arabic/Urdu UI navigation text
• "easgpdisd", "eosgpoisd" (EasyPaisa watermarks)

NON-FINANCIAL DOCUMENTS:
• If the document is NOT a financial document (address list, contact info, cashback summary, advertisement):
  → category="Other", all other fields="N/A"
• Cashback summary (no actual transaction): category="Other"

MULTILINGUAL:
• Document may be in any language — extract data correctly regardless
• Names may be in Arabic, Chinese, Hindi, French etc. — extract as-is in original script
• Amounts may use local currency symbols — preserve them

═══════════════════════════════════════════════
REMEMBER: Return ONLY the JSON object. Nothing else.
═══════════════════════════════════════════════
"""

_GROQ_KEY_INDEX = 0


def _call_groq(text: str) -> dict:
    global _GROQ_KEY_INDEX

    if not GROQ_KEYS:
        raise RuntimeError("No Groq API keys configured.")

    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user",   "content": f"Extract all financial data from this document:\n\n{text[:6000]}"},
    ]
    payload = {
        "model":       GROQ_MODEL,
        "messages":    messages,
        "temperature": 0.0,
        "max_tokens":  600,
    }

    last_error = None
    for attempt in range(len(GROQ_KEYS) * 3):
        key_idx = (_GROQ_KEY_INDEX + attempt) % len(GROQ_KEYS)
        key     = GROQ_KEYS[key_idx]

        try:
            log.info("Calling Groq (key #%d, attempt %d)…", key_idx + 1, attempt + 1)
            resp = httpx.post(
                GROQ_URL,
                headers={
                    "Authorization": f"Bearer {key}",
                    "Content-Type":  "application/json",
                },
                json=payload,
                timeout=30,
            )

            if resp.status_code == 429:
                wait = 5 * (attempt + 1)
                log.warning("Groq key #%d rate limited — waiting %ds…", key_idx + 1, wait)
                _GROQ_KEY_INDEX = (key_idx + 1) % len(GROQ_KEYS)
                last_error = "Rate limit"
                time.sleep(wait)
                continue

            if resp.status_code != 200:
                log.warning("Groq key #%d HTTP %d", key_idx + 1, resp.status_code)
                last_error = f"HTTP {resp.status_code}"
                time.sleep(2)
                continue

            raw = resp.json()["choices"][0]["message"]["content"].strip()
            raw = re.sub(r"^```[a-z]*\n?", "", raw)
            raw = re.sub(r"\n?```$",       "", raw)
            raw = raw.strip()

            result = json.loads(raw)

            # Ensure all keys present
            for k in ["category","sender_name","receiver_name","amount",
                       "date","reference_number","bank_name"]:
                if k not in result or not str(result[k]).strip():
                    result[k] = "N/A"

            _GROQ_KEY_INDEX = key_idx
            result = _clean_result(result)
            log.info("✅ Extracted: %s", result)
            return result

        except json.JSONDecodeError as e:
            log.error("Groq JSON error: %s | raw: %.300s", e, raw)
            # Retry with explicit JSON reminder
            messages[1]["content"] += "\n\nIMPORTANT: Return ONLY valid JSON, nothing else."
            last_error = str(e)
            continue
        except Exception as e:
            log.error("Groq error (key #%d): %s", key_idx + 1, e)
            last_error = str(e)
            time.sleep(2)
            continue

    raise RuntimeError(f"All Groq keys failed. Last: {last_error}")


# ─────────────────────────────────────────────────────────────────────────────
# REGEX FALLBACK (if Groq completely fails)
# ─────────────────────────────────────────────────────────────────────────────
def _regex_fallback(text: str) -> dict:
    log.warning("Using regex fallback.")
    MONTH_MAP = {
        "jan":"01","feb":"02","mar":"03","apr":"04","may":"05","jun":"06",
        "jul":"07","aug":"08","sep":"09","oct":"10","nov":"11","dec":"12",
    }
    def norm_date(r):
        r = r.strip()
        if re.match(r"^\d{4}-\d{2}-\d{2}$", r): return r
        m = re.match(r"^(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{4})$", r)
        if m: return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
        m = re.match(r"^(\d{1,2})-([A-Za-z]+)-(\d{4})$", r, re.I)
        if m:
            mo = MONTH_MAP.get(m.group(2)[:3].lower())
            if mo: return f"{m.group(3)}-{mo}-{m.group(1).zfill(2)}"
        m = re.match(r"^(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})$", r, re.I)
        if m:
            mo = MONTH_MAP.get(m.group(2)[:3].lower())
            if mo: return f"{m.group(3)}-{mo}-{m.group(1).zfill(2)}"
        return r

    cat = "Document"
    for pat, lbl in [
        (r"transaction\s*successful|fund\s*transfer|funds?\s*transfer", "Fund Transfer Receipt"),
        (r"cash\s*with\s*draw", "Cash Withdrawal"),
        (r"invoice", "Invoice"),
        (r"receipt", "Receipt"),
        (r"salary|payslip", "Salary Slip"),
        (r"utility|electricity|sngpl|ssgc", "Utility Bill"),
        (r"bill\s*pay|bill has been paid", "Bill Payment"),
        (r"medical|hospital|clinic|lab", "Medical Bill"),
    ]:
        if re.search(pat, text, re.I): cat = lbl; break

    amt = "N/A"
    for pat, prefix in [
        (r"Rs\.?\s*([\d,]+(?:\.\d{1,2})?)", "Rs. "),
        (r"PKR\.?\s*([\d,]+(?:\.\d{1,2})?)", "PKR "),
        (r"USD\s*([\d,]+(?:\.\d{1,2})?)", "USD "),
        (r"EUR\s*([\d,]+(?:\.\d{1,2})?)", "EUR "),
        (r"\$([\d,]+(?:\.\d{1,2})?)", "USD "),
        (r"£([\d,]+(?:\.\d{1,2})?)", "GBP "),
        (r"€([\d,]+(?:\.\d{1,2})?)", "EUR "),
    ]:
        m = re.search(pat, text, re.I)
        if m: amt = f"{prefix}{m.group(1)}"; break

    date = "N/A"
    for pat in [
        r"(\d{4}-\d{2}-\d{2})",
        r"(\d{1,2}-(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*-\d{4})",
        r"(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{4})",
        r"(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{4})",
    ]:
        m = re.search(pat, text, re.I)
        if m: date = norm_date(m.group(1)); break

    ref = "N/A"
    m = re.search(
        r"(?:reference\s*(?:number|no|#)|tid|receipt\s*id|transaction\s*id|ref\s*#?)[:\s|/0#]*(\d+)",
        text, re.I)
    if m: ref = m.group(1).strip()

    sender = "N/A"
    for pat in [
        r"(?:from\s*account|sent\s*by|paid\s*by|from)[:\s]+([A-Za-z][A-Za-z\s]{2,50}?)(?:\n|$|\d{4}|\*{2})",
    ]:
        m = re.search(pat, text, re.I)
        if m:
            sender = re.sub(r"\s+", " ", m.group(1)).strip()[:60]
            break

    receiver = "N/A"
    for pat in [
        r"(?:transferred\s*to|sent\s*to|successfully\s*sent\s*to|beneficiary)[:\s]+([A-Za-z][A-Za-z\s]{2,50}?)(?:\n|$|\d|\*)",
        r"(?:company\s*name|biller)[:\s]+([A-Za-z][A-Za-z\s]{2,50}?)(?:\n|$)",
    ]:
        m = re.search(pat, text, re.I)
        if m:
            receiver = re.sub(r"\s+", " ", m.group(1)).strip()[:60]
            break

    return _clean_result({
        "category": cat, "sender_name": sender, "receiver_name": receiver,
        "amount": amt, "date": date, "reference_number": ref, "bank_name": "N/A",
    })


def extract_fields(text: str) -> dict:
    clean_text = _fix_ocr_text(text)
    try:
        return _call_groq(clean_text)
    except Exception as e:
        log.error("Groq failed, using regex fallback: %s", e)
        return _regex_fallback(clean_text)


# ─────────────────────────────────────────────────────────────────────────────
# SESSION REGISTRY
# ─────────────────────────────────────────────────────────────────────────────
SESSION_REGISTRY: dict[str, Path] = {}


def _session_path(session_id: str, user_id: str) -> Path:
    if session_id not in SESSION_REGISTRY:
        SESSION_REGISTRY[session_id] = OUTPUT_DIR / f"session_{user_id}_{session_id[:8]}.xlsx"
    return SESSION_REGISTRY[session_id]


# ─────────────────────────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────────────────────────
@app.get("/health")
async def health():
    return {
        "status":    "ok",
        "version":   "6.0.0",
        "ocr":       "EasyOCR (4 script-group readers) + Tesseract fallback for forms/scans",
        "ai":        f"Groq {GROQ_MODEL}",
        "groq_keys": len(GROQ_KEYS),
        "timestamp": datetime.utcnow().isoformat(),
    }


@app.post("/process-document")
async def process_document(
    request:    Request,
    file:       UploadFile = File(...),
    mode:       str        = Form(...),
    user_id:    str        = Form(...),
    language:   str        = Form("english"),
    session_id: str        = Form(...),
):
    if not file.filename:
        raise HTTPException(400, "No file provided.")

    suffix  = Path(file.filename).suffix.lower()
    allowed = {".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".docx", ".webp"}
    if suffix not in allowed:
        raise HTTPException(400, f"Unsupported file type '{suffix}'.")

    tmp_path = UPLOAD_DIR / f"{uuid.uuid4()}{suffix}"
    try:
        tmp_path.write_bytes(await file.read())
        log.info("Saved: %s (%d bytes)", tmp_path.name, tmp_path.stat().st_size)

        doc_text = run_ocr(tmp_path, language)
        log.info("OCR → %d chars:\n%s", len(doc_text), doc_text[:800])

        if not doc_text.strip():
            raise HTTPException(422, "Could not extract text. Ensure image is clear and not blurry.")

        fields = extract_fields(doc_text)

        if mode == "append":
            excel_path = _session_path(session_id, user_id)
        else:
            ts         = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            excel_path = OUTPUT_DIR / f"doc_{user_id}_{ts}_{uuid.uuid4().hex[:6]}.xlsx"

        rows_written = append_rows_to_excel(excel_path, [fields], file.filename)
        base         = str(request.base_url).rstrip("/")
        download_url = f"{base}/download/{excel_path.name}"

        log.info("Done → %s | rows=%d", excel_path.name, rows_written)
        return JSONResponse({
            "success":        True,
            "download_url":   download_url,
            "rows_extracted": rows_written,
            "file_name":      excel_path.name,
        })

    except HTTPException:
        raise
    except Exception as e:
        log.exception("Error: %s", file.filename)
        raise HTTPException(500, str(e))
    finally:
        tmp_path.unlink(missing_ok=True)


@app.delete("/session/{session_id}")
async def end_session(session_id: str):
    SESSION_REGISTRY.pop(session_id, None)
    return {"success": True, "message": f"Session {session_id} cleared."}


@app.get("/download/{filename}")
async def download_file(filename: str):
    filename = Path(filename).name
    path     = OUTPUT_DIR / filename
    if not path.exists():
        raise HTTPException(404, "File not found.")
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )